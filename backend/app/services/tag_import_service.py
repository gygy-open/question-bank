"""
Tag batch import service.

Generates the Excel import template and imports an .xlsx file of tags for a
subject. Malformed rows (empty name / bad color) are atomic: if any such row
fails, nothing is created and every error is reported so the caller can fix
the whole file and retry. Duplicate names (in-file or already existing) are
not treated as errors — they're skipped and reported as a count, the rest of
the file still imports. Categories named in the file that don't exist yet are
auto-created.
"""
import io
import re
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.crud.crud_tag import tag as crud_tag
from app.crud.crud_tag_category import tag_category as crud_tag_category
from app.models.tag import Tag
from app.models.tag_category import TagCategory
from app.schemas.tag import TagCreate, TagImportResult, TagImportRowError
from app.schemas.tag_category import TagCategoryCreate

# Template column headers (Chinese, matched exactly on import).
NAME_COL = "名称"
CATEGORY_COL = "分类"
COLOR_COL = "颜色"
REQUIRED_COLS = [NAME_COL]
ALL_COLS = [NAME_COL, CATEGORY_COL, COLOR_COL]

DEFAULT_COLOR = "#cccccc"
_HEX_COLOR_RE = re.compile(r"^#(?:[0-9A-Fa-f]{3}|[0-9A-Fa-f]{6})$")

# Example rows shipped inside the template so users see the expected format.
EXAMPLE_ROWS = [
    ["高考真题", "来源", "#4F81BD"],
    ["易错题", "", ""],
]


def generate_template() -> bytes:
    """Generate the .xlsx tag import template (headers + example rows)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "标签导入"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    ws.append(ALL_COLS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for row in EXAMPLE_ROWS:
        ws.append(row)

    ws.column_dimensions["A"].width = 20
    for idx in range(2, len(ALL_COLS) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 16

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _parse_rows(file_bytes: bytes) -> pd.DataFrame:
    """Read the .xlsx into a DataFrame, validating required columns exist."""
    df = pd.read_excel(io.BytesIO(file_bytes), dtype=str)
    missing = [c for c in REQUIRED_COLS if c not in df.columns]
    if missing:
        raise ValueError(f"缺少必需的列：{', '.join(missing)}。请使用最新导入模板。")
    return df


def _clean(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    return str(value).strip()


async def import_excel(
    db: AsyncSession,
    *,
    subject_id: int,
    file_bytes: bytes,
    user_id: Optional[int] = None,
) -> TagImportResult:
    """
    Import tags from an .xlsx file for a subject.

    Malformed rows (empty name / bad color) are atomic: a single one rejects
    the whole file and nothing is created. Duplicate names are skipped rather
    than rejected, so the rest of the file still imports. Category names not
    found for the subject are auto-created (only once the file as a whole
    passes format validation).
    """
    try:
        df = _parse_rows(file_bytes)
    except ValueError as e:
        return TagImportResult(
            status="failed", created=0, failed=0, skipped=0, total=0,
            errors=[TagImportRowError(row=0, message=str(e))],
        )

    total = len(df)

    categories_result = await db.execute(
        select(TagCategory).where(TagCategory.subject_id == subject_id)
    )
    category_by_name: Dict[str, int] = {
        c.name.strip(): c.id for c in categories_result.scalars().all()
    }

    existing_result = await db.execute(select(Tag.name).where(Tag.subject_id == subject_id))
    existing_names = {n for (n,) in existing_result.all()}

    errors: List[TagImportRowError] = []
    seen_in_file: set = set()
    parsed: List[Dict[str, Any]] = []
    categories_to_create: set = set()
    skipped = 0

    for idx, row in df.iterrows():
        excel_row = int(idx) + 2  # +1 header, +1 for 1-based
        name = _clean(row.get(NAME_COL))
        category_name = _clean(row.get(CATEGORY_COL))
        color = _clean(row.get(COLOR_COL))

        if not name:
            errors.append(TagImportRowError(row=excel_row, message="名称不能为空"))
            continue
        if name in seen_in_file or name in existing_names:
            skipped += 1
            continue

        if color and not _HEX_COLOR_RE.match(color):
            errors.append(TagImportRowError(
                row=excel_row, message=f"颜色 '{color}' 格式不正确，需为十六进制色值（如 #4F81BD）"
            ))
            continue

        if category_name and category_name not in category_by_name:
            categories_to_create.add(category_name)

        seen_in_file.add(name)
        parsed.append({"name": name, "category_name": category_name or None, "color": color or DEFAULT_COLOR})

    if errors:
        return TagImportResult(status="failed", created=0, failed=len(errors), skipped=0, total=total, errors=errors)

    # Auto-create categories referenced by the file that don't exist yet for this subject.
    for category_name in categories_to_create:
        category = await crud_tag_category.create(
            db, obj_in=TagCategoryCreate(name=category_name, subject_id=subject_id)
        )
        category_by_name[category_name] = category.id

    for item in parsed:
        category_id = category_by_name.get(item["category_name"]) if item["category_name"] else None
        await crud_tag.create(
            db,
            obj_in=TagCreate(
                subject_id=subject_id,
                name=item["name"],
                category_id=category_id,
                color=item["color"],
            ),
            user_id=user_id,
        )

    return TagImportResult(status="success", created=len(parsed), failed=0, skipped=skipped, total=total, errors=[])
