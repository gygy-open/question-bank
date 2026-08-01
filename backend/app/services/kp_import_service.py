"""
Knowledge point batch import service.

Handles Excel template generation, preflight checks (for the rebuild mode),
and importing an .xlsx file into the knowledge point tree.
"""
import asyncio
import io
import time
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.crud.crud_knowledge_point import knowledge_point as crud_kp
from app.core.vector_store import VectorStore
from app.models.knowledge_point import KnowledgePoint
from app.models.question import question_knowledge_points
from app.models.subject import Subject
from app.schemas.knowledge_point import (
    KPImportPreflight,
    KPImportResult,
    KPImportRowError,
)

# Template column headers (Chinese, matched exactly on import).
SUBJECT_COL = "学科名称"
LEVEL_COLS = ["一级目录", "二级目录", "三级目录", "四级目录", "五级目录"]
ALL_COLS = [SUBJECT_COL] + LEVEL_COLS

# Example rows shipped inside the template so users see the expected format.
EXAMPLE_ROWS = [
    ["数学", "代数", "方程", "一元一次方程", "", ""],
    ["数学", "代数", "方程", "一元二次方程", "解法", "公式法"],
    ["数学", "几何", "平面几何", "三角形", "", ""],
]


def generate_template() -> bytes:
    """Generate the .xlsx import template (headers + example rows)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "知识点导入"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    ws.append(ALL_COLS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for row in EXAMPLE_ROWS:
        ws.append(row)

    # Reasonable column widths.
    ws.column_dimensions["A"].width = 16
    for col in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


async def preflight(db: AsyncSession, subject_id: int) -> Optional[KPImportPreflight]:
    """
    Impact assessment for the rebuild mode: how many knowledge points exist and
    how many distinct questions are linked to them.
    """
    subject = await db.get(Subject, subject_id)
    if not subject:
        return None

    existing = await db.execute(
        select(func.count(KnowledgePoint.id)).where(
            KnowledgePoint.subject_id == subject_id
        )
    )
    existing_count = existing.scalar_one()

    affected = await db.execute(
        select(func.count(func.distinct(question_knowledge_points.c.question_id)))
        .select_from(question_knowledge_points)
        .join(
            KnowledgePoint,
            KnowledgePoint.id == question_knowledge_points.c.knowledge_point_id,
        )
        .where(KnowledgePoint.subject_id == subject_id)
    )
    affected_questions = affected.scalar_one()

    return KPImportPreflight(
        subject_id=subject_id,
        subject_name=subject.name,
        existing_count=existing_count,
        affected_questions=affected_questions,
    )


def _parse_rows(file_bytes: bytes) -> pd.DataFrame:
    """Read the .xlsx into a DataFrame, validating required columns exist."""
    df = pd.read_excel(io.BytesIO(file_bytes), dtype=str)
    missing = [c for c in ALL_COLS if c not in df.columns]
    if missing:
        raise ValueError(f"缺少必需的列：{', '.join(missing)}。请使用最新导入模板。")
    return df


def _clean(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    return str(value).strip()


async def import_excel(
    db: AsyncSession,
    *,
    file_bytes: bytes,
    mode: str,
    user_id: Optional[int] = None,
) -> KPImportResult:
    """
    Import knowledge points from an .xlsx file.

    mode = "incremental": skip existing nodes, create new ones.
    mode = "rebuild": clear the target subject(s) first, then import.
    """
    started = time.perf_counter()

    try:
        df = _parse_rows(file_bytes)
    except ValueError as e:
        return KPImportResult(status="failed", mode=mode, errors=[KPImportRowError(row=0, message=str(e))])

    # Cache subjects by name and slug for fast lookup.
    subjects_result = await db.execute(select(Subject))
    subjects = subjects_result.scalars().all()
    subject_by_key: Dict[str, Subject] = {}
    for s in subjects:
        subject_by_key[s.name.strip()] = s
        subject_by_key[s.slug.strip()] = s

    errors: List[KPImportRowError] = []
    # Parsed valid rows: (subject, [path parts]). Also track which subjects are touched.
    parsed: List[Tuple[Subject, List[str]]] = []
    touched_subject_ids: set = set()

    for idx, row in df.iterrows():
        excel_row = int(idx) + 2  # +1 header, +1 for 1-based
        subject_name = _clean(row.get(SUBJECT_COL))
        if not subject_name:
            errors.append(KPImportRowError(row=excel_row, message="学科名称不能为空"))
            continue
        subject = subject_by_key.get(subject_name)
        if not subject:
            errors.append(KPImportRowError(
                row=excel_row, message=f"学科 '{subject_name}' 不存在，请先创建该学科"
            ))
            continue

        path = [p for p in (_clean(row.get(c)) for c in LEVEL_COLS) if p]
        if not path:
            errors.append(KPImportRowError(row=excel_row, message="一级目录不能为空"))
            continue

        parsed.append((subject, path))
        touched_subject_ids.add(subject.id)

    subject_name_out = None
    if len(touched_subject_ids) == 1:
        only_id = next(iter(touched_subject_ids))
        subject_name_out = next(s.name for s in subjects if s.id == only_id)

    # Rebuild mode: clear touched subjects up-front.
    if mode == "rebuild":
        for sid in touched_subject_ids:
            await crud_kp.clear_by_subject(db, subject_id=sid)

    # Build an in-memory cache of existing nodes keyed by (subject_id, parent_id, name).
    # After a rebuild the subject is empty, so this naturally starts fresh.
    cache: Dict[Tuple[int, Optional[int], str], int] = {}
    used_slugs_by_subject: Dict[int, set] = {}
    for sid in touched_subject_ids:
        existing = await crud_kp.get_by_subject(db, subject_id=sid, limit=None)
        used = set()
        for kp in existing:
            cache[(kp.subject_id, kp.parent_id, kp.name)] = kp.id
            used.add(kp.slug)
        used_slugs_by_subject[sid] = used

    created = 0
    skipped = 0
    new_ids: List[int] = []

    for subject, path in parsed:
        parent_id: Optional[int] = None
        row_created_any = False
        for name in path:
            key = (subject.id, parent_id, name)
            if key in cache:
                parent_id = cache[key]
                continue
            slug = await crud_kp.make_unique_slug(
                db, subject.id, name, used_slugs_by_subject[subject.id]
            )
            created_objs = await crud_kp.create_batch(
                db,
                nodes=[{
                    "name": name,
                    "slug": slug,
                    "subject_id": subject.id,
                    "parent_id": parent_id,
                }],
                user_id=user_id,
            )
            new_kp = created_objs[0]
            cache[key] = new_kp.id
            new_ids.append(new_kp.id)
            parent_id = new_kp.id
            created += 1
            row_created_any = True
        if not row_created_any:
            skipped += 1

    # Batch vectorization (only if embedding configured; otherwise deferred).
    vector_synced = False
    if new_ids and VectorStore.is_available():
        try:
            items = []
            for kp_id in new_ids:
                kp = await crud_kp.get(db, kp_id)
                if not kp:
                    continue
                path_text = await crud_kp._build_path_text(db, kp)
                items.append({
                    "id": kp.id,
                    "text": path_text,
                    "metadata": {
                        "id": kp.id,
                        "subject_id": kp.subject_id,
                        "name": kp.name,
                        "slug": kp.slug,
                    },
                })
            await asyncio.to_thread(VectorStore.upsert_knowledge_points_batch, items)
            vector_synced = True
        except Exception as e:
            print(f"Batch vectorization failed (can reindex manually later): {e}")

    total = len(parsed) + len(errors)
    if errors and created == 0 and skipped == 0:
        status = "failed"
    elif errors:
        status = "partial"
    else:
        status = "success"

    return KPImportResult(
        status=status,
        subject_name=subject_name_out,
        mode=mode,
        created=created,
        skipped=skipped,
        failed=len(errors),
        total=total,
        duration=round(time.perf_counter() - started, 2),
        vector_synced=vector_synced,
        errors=errors,
    )
