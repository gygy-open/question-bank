"""
User batch import service.

Generates the Excel import template and imports an .xlsx file of users,
validating each row and returning a per-row result.
"""
import io
from typing import Any, List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.crud.crud_user import user as crud_user
from app.models.user import User
from app.schemas.user import UserCreate, UserImportResult, UserImportRowError

# Template column headers (Chinese, matched exactly on import).
USERNAME_COL = "用户名"
FULL_NAME_COL = "姓名"
PASSWORD_COL = "密码"
ROLE_COL = "管理员"  # optional: 是/true/1 => superuser
REQUIRED_COLS = [USERNAME_COL, FULL_NAME_COL, PASSWORD_COL]
ALL_COLS = REQUIRED_COLS + [ROLE_COL]

MIN_PASSWORD_LENGTH = 6

# Example rows shipped inside the template so users see the expected format.
EXAMPLE_ROWS = [
    ["teacher01", "张三", "changeme123", ""],
    ["admin02", "李四", "changeme123", "是"],
]

_TRUE_VALUES = {"是", "true", "1", "yes", "y", "admin", "管理员"}


def generate_template() -> bytes:
    """Generate the .xlsx user import template (headers + example rows)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "用户导入"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    ws.append(ALL_COLS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for row in EXAMPLE_ROWS:
        ws.append(row)

    ws.column_dimensions["A"].width = 18
    for idx in range(2, len(ALL_COLS) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 16

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _parse_rows(file_bytes: bytes) -> pd.DataFrame:
    """Read the .xlsx into a DataFrame, validating required columns exist."""
    df = pd.read_excel(io.BytesIO(file_bytes), dtype=str)
    missing = [c for c in REQUIRED_COLS if c not in df.columns]
    if missing:
        raise ValueError(f"缺少必需的列：{', '.join(missing)}。请使用最新导入模板。")
    return df


def _clean(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    return str(value).strip()


def _is_truthy(value: str) -> bool:
    return value.strip().lower() in _TRUE_VALUES


async def import_excel(
    db: AsyncSession,
    *,
    file_bytes: bytes,
) -> UserImportResult:
    """Import users from an .xlsx file, validating each row."""
    try:
        df = _parse_rows(file_bytes)
    except ValueError as e:
        return UserImportResult(
            status="failed", errors=[UserImportRowError(row=0, message=str(e))]
        )

    total = len(df)

    # Cache existing usernames for fast duplicate detection.
    existing_result = await db.execute(select(User.username))
    existing_usernames = {u for (u,) in existing_result.all()}

    errors: List[UserImportRowError] = []
    seen_in_file: set = set()
    created = 0

    for idx, row in df.iterrows():
        excel_row = int(idx) + 2  # +1 header, +1 for 1-based
        username = _clean(row.get(USERNAME_COL))
        full_name = _clean(row.get(FULL_NAME_COL))
        password = _clean(row.get(PASSWORD_COL))
        has_role_col = ROLE_COL in df.columns
        is_superuser = _is_truthy(_clean(row.get(ROLE_COL))) if has_role_col else False

        if not username:
            errors.append(UserImportRowError(row=excel_row, message="用户名不能为空"))
            continue
        if not full_name:
            errors.append(UserImportRowError(row=excel_row, message="姓名不能为空"))
            continue
        if not password:
            errors.append(UserImportRowError(row=excel_row, message="密码不能为空"))
            continue
        if len(password) < MIN_PASSWORD_LENGTH:
            errors.append(UserImportRowError(
                row=excel_row, message=f"密码长度不能少于 {MIN_PASSWORD_LENGTH} 位"
            ))
            continue
        if username in seen_in_file:
            errors.append(UserImportRowError(
                row=excel_row, message=f"用户名 '{username}' 在文件中重复"
            ))
            continue
        if username in existing_usernames:
            errors.append(UserImportRowError(
                row=excel_row, message=f"用户名 '{username}' 已存在"
            ))
            continue

        await crud_user.create(
            db,
            obj_in=UserCreate(
                username=username,
                full_name=full_name,
                password=password,
                is_active=True,
                is_superuser=is_superuser,
            ),
        )
        seen_in_file.add(username)
        existing_usernames.add(username)
        created += 1

    failed = len(errors)
    if created == 0 and failed > 0:
        status = "failed"
    elif failed > 0:
        status = "partial"
    else:
        status = "success"

    return UserImportResult(
        status=status, created=created, failed=failed, total=total, errors=errors
    )
